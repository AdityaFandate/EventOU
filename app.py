from datetime import datetime, timedelta
import uuid
import os
import secrets

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    jsonify,
    flash,
    send_file,
)
import io
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import check_password_hash, generate_password_hash
import qrcode


app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///event_crowd.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = "change-this-secret-key"
app.config["QR_FOLDER"] = os.path.join("static", "qr")

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    roll_number = db.Column(db.String(50), nullable=True)  # College roll number
    role = db.Column(db.String(20), nullable=False, default="student")  # student, admin, faculty_coordinator, volunteer, security, host, canteen_owner
    is_guest = db.Column(db.Boolean, default=False)  # For temporary guest/alumni passes

    events = db.relationship("Event", back_populates="host", lazy=True)
    tickets = db.relationship("Ticket", back_populates="student", lazy=True)

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)


class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    location = db.Column(db.String(200), nullable=False)
    start_time = db.Column(db.DateTime, nullable=False)
    end_time = db.Column(db.DateTime, nullable=False)
    max_capacity = db.Column(db.Integer, nullable=False)
    safe_threshold = db.Column(db.Float, default=0.7)  # 70%
    warning_threshold = db.Column(db.Float, default=0.9)  # 90%
    current_count = db.Column(db.Integer, default=0)
    entry_blocked = db.Column(db.Boolean, default=False)  # Admin can block entry
    host_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    event_type = db.Column(db.String(50), default="general")  # cultural, technical, sports, seminar, convocation
    poster_filename = db.Column(db.String(255), nullable=True)
    
    # New: Pricing fields
    is_paid = db.Column(db.Boolean, default=False)
    price = db.Column(db.Float, default=0.0)
    
    # New: Group event field
    is_group_event = db.Column(db.Boolean, default=False)
    group_size = db.Column(db.Integer, default=1)  # Max members per group registration

    host = db.relationship("User", back_populates="events")
    tickets = db.relationship("Ticket", back_populates="event", lazy=True, cascade="all, delete-orphan")
    zones = db.relationship("Zone", back_populates="event", lazy=True, cascade="all, delete-orphan")
    time_slots = db.relationship("TimeSlot", back_populates="event", lazy=True, cascade="all, delete-orphan")
    gates = db.relationship("Gate", back_populates="event", lazy=True, cascade="all, delete-orphan")

    @property
    def occupancy_ratio(self) -> float:
        if self.max_capacity <= 0:
            return 0.0
        return self.current_count / self.max_capacity

    @property
    def status(self) -> str:
        if self.entry_blocked:
            return "blocked"
        ratio = self.occupancy_ratio
        if ratio < self.safe_threshold:
            return "normal"
        if ratio < self.warning_threshold:
            return "warning"
        return "critical"

    @property
    def status_message(self) -> str:
        status = self.status
        if status == "blocked":
            return "Entry is currently blocked by admin."
        if status == "normal":
            return "Crowd level is within safe limits."
        if status == "warning":
            return "Crowd is approaching the upper safe limit. Monitor entrances and exits closely."
        return "Critical crowd level! Consider pausing entry and directing people to exits."


class Zone(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # Stage Area, Seating Area, Food Stalls, etc.
    max_capacity = db.Column(db.Integer, nullable=False)
    current_count = db.Column(db.Integer, default=0)
    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=False)

    event = db.relationship("Event", back_populates="zones")

    @property
    def occupancy_ratio(self) -> float:
        if self.max_capacity <= 0:
            return 0.0
        return self.current_count / self.max_capacity


class Gate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # Gate A, Gate B, Main Entrance, etc.
    entry_count = db.Column(db.Integer, default=0)
    exit_count = db.Column(db.Integer, default=0)
    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=False)

    event = db.relationship("Event", back_populates="gates")

    @property
    def net_count(self) -> int:
        return self.entry_count - self.exit_count


class TimeSlot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    start_time = db.Column(db.DateTime, nullable=False)
    end_time = db.Column(db.DateTime, nullable=False)
    max_participants = db.Column(db.Integer, nullable=False)
    current_participants = db.Column(db.Integer, default=0)
    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=False)

    event = db.relationship("Event", back_populates="time_slots")
    tickets = db.relationship("Ticket", back_populates="time_slot", lazy=True)

    @property
    def is_full(self) -> bool:
        return self.current_participants >= self.max_participants

    @property
    def is_active(self) -> bool:
        now = datetime.now()
        return self.start_time <= now <= self.end_time


class Ticket(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    qr_token = db.Column(db.String(64), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    paid = db.Column(db.Boolean, default=False)
    checked_in = db.Column(db.Boolean, default=False)
    checked_in_at = db.Column(db.DateTime, nullable=True)
    gate_id = db.Column(db.Integer, db.ForeignKey("gate.id"), nullable=True)
    
    # New: Group details
    guest_name = db.Column(db.String(100), nullable=True)
    team_name = db.Column(db.String(100), nullable=True)
    member_details = db.Column(db.Text, nullable=True)  # Store JSON or text details of group members

    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    time_slot_id = db.Column(db.Integer, db.ForeignKey("time_slot.id"), nullable=True)

    event = db.relationship("Event", back_populates="tickets")
    student = db.relationship("User", back_populates="tickets")
    time_slot = db.relationship("TimeSlot", back_populates="tickets")
    gate = db.relationship("Gate")


class Announcement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    is_active = db.Column(db.Boolean, default=True)
    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=True)  # None = global announcement
    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)

    event = db.relationship("Event")
    created_by = db.relationship("User")


class FoodCouponRedemption(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    qr_data = db.Column(db.String(255), nullable=False)
    redeemed_at = db.Column(db.DateTime, default=datetime.now)
    redeemed_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    
    # Linked fields for better reporting
    coupon_id = db.Column(db.Integer, db.ForeignKey("food_coupon.id"), nullable=True)
    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=True)

    redeemed_by = db.relationship("User", foreign_keys=[redeemed_by_id])
    coupon = db.relationship("FoodCoupon")
    event = db.relationship("Event")


class FoodCoupon(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    coupon_name = db.Column(db.String(100), nullable=False)
    food_item = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, default=1)
    qr_code = db.Column(db.String(200), unique=True, nullable=False)
    status = db.Column(db.String(20), default="Pending")
    used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    expiry_date = db.Column(db.DateTime, nullable=True)
    used_at = db.Column(db.DateTime, nullable=True)
    created_by = db.Column(db.String(50), nullable=True)
    meal_type = db.Column(db.String(20), default="other")  # breakfast, lunch, dinner, other
    
    # New fields to link coupons to events, students, and specific tickets
    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=True)
    student_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey("ticket.id"), nullable=True)

    event = db.relationship("Event")
    student = db.relationship("User", foreign_keys=[student_id])
    ticket = db.relationship("Ticket")

@login_manager.user_loader
def load_user(user_id: str):
    return db.session.get(User, int(user_id))


@app.context_processor
def inject_now():
    return {"now": datetime.now()}


def require_role(*allowed_roles):
    """Decorator to check user role"""
    def decorator(f):
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                flash("Please log in to access this page.", "danger")
                return redirect(url_for("login"))
            if current_user.role not in allowed_roles:
                flash("You don't have permission to access this page.", "danger")
                return redirect(url_for("index"))
            return f(*args, **kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator


@app.route("/")
def index():
    if current_user.is_authenticated:
        # Redirect based on role
        if current_user.role == "admin":
            return redirect(url_for("admin_dashboard"))
        elif current_user.role == "faculty_coordinator":
            return redirect(url_for("faculty_dashboard"))
        elif current_user.role == "host":
            return redirect(url_for("host_dashboard"))
        elif current_user.role == "volunteer":
            return redirect(url_for("volunteer_scanner"))
        elif current_user.role == "security":
            return redirect(url_for("security_dashboard"))
        elif current_user.role == "canteen_owner":
            return redirect(url_for("canteen_dashboard"))
        else:
            # For students or any other role, show event listing or a student dashboard
            return redirect(url_for("event_listing"))

    # Modern marketing-style landing page with real data for guests
    active_events = Event.query.count()
    total_capacity = db.session.query(db.func.sum(Event.max_capacity)).scalar() or 0
    # User roles count (distinct)
    role_count = 7 # student, admin, faculty_coordinator, volunteer, security, host, canteen_owner
    
    # Get top 3 upcoming or ongoing events for the live overview card
    now = datetime.now()
    live_events = Event.query.filter(Event.end_time >= now).order_by(Event.start_time.asc()).limit(3).all()
    
    # Total tickets scanned (placeholder logic or real)
    tickets_scanned = Ticket.query.filter_by(checked_in=True).count()
    
    return render_template("landing.html", 
                           active_events=active_events, 
                           total_capacity=total_capacity,
                           role_count=role_count,
                           live_events=live_events,
                           tickets_scanned=tickets_scanned)


@app.route("/events")
def event_listing():
    now = datetime.now()
    # Only show events that haven't ended yet
    events = Event.query.filter(Event.end_time >= now).order_by(Event.start_time.asc()).all()
    return render_template("event_listing.html", events=events, now=now)


@app.route("/events/<int:event_id>")
def event_detail(event_id: int):
    event = Event.query.get_or_404(event_id)
    zones = Zone.query.filter_by(event_id=event_id).all()
    gates = Gate.query.filter_by(event_id=event_id).all()
    time_slots = TimeSlot.query.filter_by(event_id=event_id).order_by(TimeSlot.start_time).all()
    announcements = Announcement.query.filter_by(event_id=event_id, is_active=True).order_by(Announcement.created_at.desc()).all()
    
    user_tickets = []
    coupons = []
    if current_user.is_authenticated and current_user.role == "student":
        user_tickets = Ticket.query.filter_by(event_id=event_id, student_id=current_user.id).all()
        if user_tickets:
            coupons = FoodCoupon.query.filter_by(event_id=event_id, student_id=current_user.id).all()
    
    return render_template("event_detail.html", event=event, zones=zones, gates=gates, time_slots=time_slots, announcements=announcements, user_tickets=user_tickets, coupons=coupons)


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        roll_number = request.form.get("roll_number", "").strip()
        role = request.form.get("role") or "student"
        is_guest = request.form.get("is_guest") == "on"

        if not name or not email or not password:
            flash("All fields are required.", "danger")
            return redirect(url_for("register"))

        # Validate college email for students (optional check)
        if role == "student" and not is_guest:
            if "@" not in email or "." not in email:
                flash("Please use a valid college email address.", "danger")
                return redirect(url_for("register"))

        if role not in {"student", "host", "faculty_coordinator", "volunteer", "security", "canteen_owner"}:
            role = "student"

        existing = User.query.filter_by(email=email).first()
        if existing:
            flash("Email already registered. Please log in.", "warning")
            return redirect(url_for("login"))

        user = User(name=name, email=email, role=role, roll_number=roll_number if roll_number else None, is_guest=is_guest)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        flash("Registration successful. You can now log in.", "success")
        return redirect(url_for("login"))

    return render_template("auth_register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = User.query.filter_by(email=email).first()
        if not user or not user.check_password(password):
            flash("Invalid email or password.", "danger")
            return redirect(url_for("login"))

        login_user(user)
        flash("Logged in successfully.", "success")
        
        # Redirect based on role
        if user.role == "admin":
            return redirect(url_for("admin_dashboard"))
        elif user.role == "faculty_coordinator":
            return redirect(url_for("faculty_dashboard"))
        elif user.role == "host":
            return redirect(url_for("host_dashboard"))
        elif user.role == "volunteer":
            return redirect(url_for("volunteer_scanner"))
        elif user.role == "security":
            return redirect(url_for("security_dashboard"))
        elif user.role == "canteen_owner":
            return redirect(url_for("canteen_dashboard"))
        else:
            return redirect(url_for("index"))

    return render_template("auth_login.html")


@app.route("/events/<int:event_id>/delete", methods=["POST"])
@login_required
@require_role("admin", "host")
def delete_event(event_id: int):
    event = Event.query.get_or_404(event_id)
    
    # Check permission
    if current_user.role != "admin" and event.host_id != current_user.id:
        flash("You do not have permission to delete this event.", "danger")
        return redirect(url_for("index"))
    
    try:
        db.session.delete(event)
        db.session.commit()
        flash(f"Event '{event.name}' has been deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting event: {str(e)}", "danger")
    
    return redirect(request.referrer or url_for("index"))


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = User.query.filter_by(email=email).first()
        if user:
            # In a real app, send email. Here, redirect to reset page for demonstration.
            flash("User found. You can now reset your password.", "success")
            return redirect(url_for("reset_password", email=email))
        else:
            flash("No account found with that email address.", "danger")
    return render_template("forgot_password.html")


@app.route("/forgot-username", methods=["GET", "POST"])
def forgot_username():
    if request.method == "POST":
        roll_number = request.form.get("roll_number", "").strip()
        name = request.form.get("name", "").strip()
        user = User.query.filter_by(roll_number=roll_number, name=name).first()
        if user:
            flash(f"Your registered email (username) is: {user.email}", "info")
        else:
            flash("No account found with those details. Please check your Roll Number and Name.", "danger")
    return render_template("forgot_username.html")


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    email = request.args.get("email") or request.form.get("email")
    if not email:
        flash("Invalid reset request.", "danger")
        return redirect(url_for("forgot_password"))
        
    if request.method == "POST":
        new_password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")
        
        if not new_password or new_password != confirm_password:
            flash("Passwords do not match.", "danger")
            return render_template("reset_password.html", email=email)
            
        user = User.query.filter_by(email=email).first()
        if user:
            user.password = generate_password_hash(new_password)
            db.session.commit()
            flash("Your password has been reset successfully. Please login with your new password.", "success")
            return redirect(url_for("login"))
        else:
            flash("User not found.", "danger")
            return redirect(url_for("forgot_password"))
            
    return render_template("reset_password.html", email=email)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("index"))


@app.route("/admin/dashboard")
@login_required
@require_role("admin")
def admin_dashboard():
    now = datetime.now()
    # Filter only active/upcoming events
    events = Event.query.filter(Event.end_time >= now).order_by(Event.start_time.desc()).all()
    users = User.query.all()
    total_tickets = Ticket.query.count()
    active_announcements = Announcement.query.filter_by(is_active=True).count()
    
    # New: Fetch Entry Logs and Coupon History for Admin
    entry_logs = Ticket.query.filter(Ticket.checked_in == True).order_by(Ticket.checked_in_at.desc()).all()
    coupon_history = FoodCouponRedemption.query.order_by(FoodCouponRedemption.redeemed_at.desc()).all()
    
    return render_template(
        "admin_dashboard.html", 
        events=events, 
        users=users, 
        total_tickets=total_tickets, 
        active_announcements=active_announcements,
        entry_logs=entry_logs,
        coupon_history=coupon_history,
        now=now
    )


@app.route("/faculty/dashboard")
@login_required
@require_role("faculty_coordinator", "admin")
def faculty_dashboard():
    now = datetime.now()
    # Filter only active/upcoming events
    events = Event.query.filter(Event.end_time >= now).order_by(Event.start_time.desc()).all()
    return render_template("faculty_dashboard.html", events=events, now=now)


@app.route("/security/dashboard")
@login_required
@require_role("security", "admin", "faculty_coordinator")
def security_dashboard():
    now = datetime.now()
    # Filter only active/upcoming events
    events = Event.query.filter(Event.end_time >= now).order_by(Event.start_time.desc()).all()
    alerts = []
    for event in events:
        if event.status in {"warning", "critical"}:
            alerts.append(event)
    total_coupons = FoodCouponRedemption.query.count()
    today = now.date()
    today_coupons = FoodCouponRedemption.query.filter(
        db.func.date(FoodCouponRedemption.redeemed_at) == today
    ).count()
    return render_template(
        "security_dashboard.html",
        events=events,
        alerts=alerts,
        total_coupons=total_coupons,
        today_coupons=today_coupons,
        now=now
    )


@app.route("/canteen/dashboard")
@login_required
@require_role("canteen_owner", "admin")
def canteen_dashboard():
    total_coupons = FoodCouponRedemption.query.count()
    today = datetime.now().date()
    today_coupons = FoodCouponRedemption.query.filter(
        db.func.date(FoodCouponRedemption.redeemed_at) == today
    ).count()

    by_user = (
        db.session.query(User.name, db.func.count(FoodCouponRedemption.id))
        .join(FoodCouponRedemption, FoodCouponRedemption.redeemed_by_id == User.id)
        .group_by(User.id)
        .all()
    )

    return render_template(
        "canteen_dashboard.html",
        total_coupons=total_coupons,
        today_coupons=today_coupons,
        by_user=by_user,
    )


@app.route("/host/dashboard")
@login_required
@require_role("host", "admin")
def host_dashboard():
    now = datetime.now()
    # Events hosted by this user (or all events for admin)
    # Filter only active/upcoming events
    if current_user.role == "admin":
        events = Event.query.filter(Event.end_time >= now).order_by(Event.start_time.desc()).all()
    else:
        events = (
            Event.query.filter(Event.host_id == current_user.id, Event.end_time >= now)
            .order_by(Event.start_time.desc())
            .all()
        )

    tickets_by_event = {
        e.id: Ticket.query.filter_by(event_id=e.id).count() for e in events
    }
    gates_by_event = {
        e.id: Gate.query.filter_by(event_id=e.id).all() for e in events
    }
    zones_by_event = {
        e.id: Zone.query.filter_by(event_id=e.id).all() for e in events
    }

    if current_user.role == "admin":
        entry_logs = (
            Ticket.query.filter(Ticket.checked_in.is_(True))
            .order_by(Ticket.checked_in_at.desc())
            .limit(10)
            .all()
        )
    else:
        entry_logs = (
            Ticket.query.join(Event)
            .filter(Ticket.checked_in.is_(True), Event.host_id == current_user.id)
            .order_by(Ticket.checked_in_at.desc())
            .limit(10)
            .all()
        )

    faculty_users = User.query.filter_by(role="faculty_coordinator").all()
    volunteer_users = User.query.filter_by(role="volunteer").all()

    total_coupons = FoodCoupon.query.filter_by(used=True).count()
    coupons = (
        FoodCoupon.query.order_by(FoodCoupon.created_at.desc())
        .limit(5)
        .all()
    )

    # New: Fetch Pass Generation History (Enrollments) for Host
    if current_user.role == "admin":
        pass_history = Ticket.query.order_by(Ticket.created_at.desc()).all()
    else:
        pass_history = Ticket.query.join(Event).filter(Event.host_id == current_user.id).order_by(Ticket.created_at.desc()).all()

    return render_template(
        "host_dashboard.html",
        events=events,
        tickets_by_event=tickets_by_event,
        gates_by_event=gates_by_event,
        zones_by_event=zones_by_event,
        entry_logs=entry_logs,
        faculty_users=faculty_users,
        volunteer_users=volunteer_users,
        total_coupons=total_coupons,
        coupons=coupons,
        pass_history=pass_history,
        now=now
    )


@app.route("/create_coupon", methods=["GET", "POST"])
@login_required
@require_role("host", "admin")
def create_coupon():
    if request.method == "POST":
        coupon_name = request.form.get("coupon_name", "").strip()
        food_item = request.form.get("food_item", "").strip()
        quantity_raw = request.form.get("quantity", "1")
        meal_type = request.form.get("meal_type", "other")
        event_id = request.form.get("event_id")
        student_id = request.form.get("student_id")

        try:
            quantity = int(quantity_raw)
        except ValueError:
            quantity = 1

        if not coupon_name or not food_item or quantity <= 0:
            flash("Please provide valid coupon details.", "danger")
            return redirect(url_for("create_coupon"))

        unique_code = str(uuid.uuid4())
        expiry = datetime.now() + timedelta(hours=4)
        
        qr_folder = os.path.join(app.root_path, "static", "qrcodes")
        os.makedirs(qr_folder, exist_ok=True)

        if event_id and student_id:
            # Single creation for a specific student, but check if they are enrolled in the event
            ticket = Ticket.query.filter_by(event_id=int(event_id), student_id=int(student_id)).first()
            if not ticket:
                flash("Selected student is not enrolled in this event.", "danger")
                return redirect(url_for("create_coupon"))
            
            coupon = FoodCoupon(
                coupon_name=coupon_name,
                food_item=food_item,
                quantity=quantity,
                created_by=current_user.email if current_user.is_authenticated else None,
                qr_code=unique_code,
                expiry_date=expiry,
                event_id=int(event_id),
                student_id=int(student_id),
                meal_type=meal_type
            )

            db.session.add(coupon)
            db.session.commit()

            img = qrcode.make(unique_code)
            img_path = os.path.join(qr_folder, unique_code + ".png")
            img.save(img_path)

            flash(f"Food coupon created for student: {coupon.student.name}", "success")

        elif event_id and not student_id:
            # Bulk creation for all tickets enrolled in this event (one coupon per ticket)
            enrolled_tickets = Ticket.query.filter_by(event_id=int(event_id)).all()
            if not enrolled_tickets:
                flash("No students are currently enrolled in this event.", "warning")
                return redirect(url_for("create_coupon"))
            
            created_count = 0
            for ticket in enrolled_tickets:
                s_code = str(uuid.uuid4())
                coupon = FoodCoupon(
                    coupon_name=coupon_name,
                    food_item=food_item,
                    quantity=quantity,
                    created_by=current_user.email if current_user.is_authenticated else None,
                    qr_code=s_code,
                    expiry_date=expiry,
                    event_id=int(event_id),
                    student_id=ticket.student_id,
                    ticket_id=ticket.id,  # Link to specific pass
                    meal_type=meal_type
                )
                db.session.add(coupon)
                
                img = qrcode.make(s_code)
                img_path = os.path.join(qr_folder, s_code + ".png")
                img.save(img_path)
                created_count += 1
            
            db.session.commit()
            flash(f"Successfully created and assigned {created_count} food coupons to all event passes.", "success")

        else:
            # General coupon (not tied to event or specific student)
            coupon = FoodCoupon(
                coupon_name=coupon_name,
                food_item=food_item,
                quantity=quantity,
                created_by=current_user.email if current_user.is_authenticated else None,
                qr_code=unique_code,
                expiry_date=expiry,
                event_id=None,
                student_id=None,
                meal_type=meal_type
            )

            db.session.add(coupon)
            db.session.commit()

            img = qrcode.make(unique_code)
            img_path = os.path.join(qr_folder, unique_code + ".png")
            img.save(img_path)

            flash("General food coupon created successfully.", "success")

    # Get events and students for the form
    if current_user.role == "admin":
        events = Event.query.all()
    else:
        events = Event.query.filter_by(host_id=current_user.id).all()
    
    students = User.query.filter_by(role="student").all()
    return render_template("create_coupon.html", events=events, students=students)


@app.route("/coupon_scanner")
@login_required
@require_role("canteen_owner", "admin")
def coupon_scanner():
    return render_template("coupon_scanner.html")


@app.route("/coupons/<int:coupon_id>")
@login_required
def view_coupon(coupon_id: int):
    coupon = FoodCoupon.query.get_or_404(coupon_id)
    
    # Allow admin and canteen owner to view all coupons
    if current_user.role in {"admin", "canteen_owner"}:
        return render_template("my_coupon.html", coupon=coupon)
    
    # Security check for students:
    if current_user.role == "student":
        # 1. If assigned to a student, must be the current user
        if coupon.student_id and coupon.student_id != current_user.id:
            flash("You are not authorized to view this coupon.", "danger")
            return redirect(url_for("index"))
            
        # 2. If tied to an event, student must be enrolled
        if coupon.event_id:
            ticket = Ticket.query.filter_by(event_id=coupon.event_id, student_id=current_user.id).first()
            if not ticket:
                flash("This coupon is restricted to event participants.", "danger")
                return redirect(url_for("index"))
    else:
        # Non-student roles (e.g. host) can't view individual coupons unless they are admin/canteen
        flash("You are not authorized to view this coupon.", "danger")
        return redirect(url_for("index"))

    return render_template("my_coupon.html", coupon=coupon)


@app.route("/verify_coupon/<code>")
def verify_coupon(code: str):
    coupon = FoodCoupon.query.filter_by(qr_code=code).first()

    if not coupon:
        message = "Invalid Coupon"
    elif coupon.used:
        message = "Coupon Already Used"
    elif coupon.expiry_date and datetime.now() > coupon.expiry_date:
        message = "Coupon Expired"
    elif not coupon.student_id:
        message = "General coupon cannot be redeemed this way."
    else:
        # Time-based validation for meal types
        if coupon.meal_type != "other":
            now = datetime.now()
            hour = now.hour
            minute = now.minute
            current_time_float = hour + minute / 60.0
            
            meal_windows = {
                "breakfast": {"start": 7.0, "end": 10.5, "label": "Breakfast (7:00 AM - 10:30 AM)"},
                "lunch": {"start": 12.0, "end": 15.5, "label": "Lunch (12:00 PM - 3:30 PM)"},
                "snacks": {"start": 16.5, "end": 18.5, "label": "Evening Snacks (4:30 PM - 6:30 PM)"},
                "dinner": {"start": 19.0, "end": 22.5, "label": "Dinner (7:00 PM - 10:30 PM)"}
            }
            
            if coupon.meal_type in meal_windows:
                window = meal_windows[coupon.meal_type]
                if current_time_float < window["start"] or current_time_float > window["end"]:
                    return render_template("coupon_result.html", 
                                        message=f"Invalid Time. {window['label']} coupons can only be scanned during their respective time.", 
                                        coupon=coupon)

        # Check if student is enrolled in the event (if applicable)
        if coupon.event_id:
            ticket = Ticket.query.filter_by(event_id=coupon.event_id, student_id=coupon.student_id).first()
            if not ticket:
                return render_template("coupon_result.html", message="Student is not enrolled in this event.", coupon=coupon)

        coupon.used = True
        coupon.used_at = datetime.now()
        coupon.status = "Used"
        
        # Log redemption with event and coupon details
        redemption = FoodCouponRedemption(
            qr_data=code,
            redeemed_by_id=current_user.id if current_user.is_authenticated else None,
            coupon_id=coupon.id,
            event_id=coupon.event_id
        )
        db.session.add(redemption)
        db.session.commit()
        message = "Coupon Valid – Food Served"

    return render_template("coupon_result.html", message=message, coupon=coupon)


@app.route("/coupon_report")
@login_required
@require_role("admin", "host")
def coupon_report():
    coupons = FoodCoupon.query.order_by(FoodCoupon.created_at.desc()).all()
    return render_template("coupon_report.html", coupons=coupons)


@app.route("/volunteer/scanner")
@login_required
@require_role("volunteer", "security", "host", "admin")
def volunteer_scanner():
    events = Event.query.filter(Event.end_time >= datetime.now()).order_by(Event.start_time).all()
    return render_template("volunteer_scanner.html", events=events)


@app.route("/events/create", methods=["POST"])
@login_required
def create_event():
    try:
        name = request.form.get("name", "").strip()
        location = request.form.get("location", "").strip()
        start_time_str = request.form.get("start_time")
        end_time_str = request.form.get("end_time")
        max_capacity = int(request.form.get("max_capacity") or 0)
        safe_threshold = float(request.form.get("safe_threshold") or 70) / 100.0
        warning_threshold = float(request.form.get("warning_threshold") or 90) / 100.0
        event_type = request.form.get("event_type") or "general"
        is_paid = request.form.get("is_paid") == "1"
        price = float(request.form.get("price") or 0.0) if is_paid else 0.0
        is_group_event = request.form.get("is_group_event") == "1"
        group_size = int(request.form.get("group_size") or 1)

        if not name or not location or max_capacity <= 0:
            raise ValueError("Please fill all fields and use a positive capacity.")

        start_time = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M")
        end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M")

        if end_time <= start_time:
            raise ValueError("End time must be after start time.")

        if not (0 < safe_threshold < warning_threshold <= 1):
            raise ValueError("Thresholds must be between 0–100 and safe < warning.")

        event = Event(
            name=name,
            location=location,
            start_time=start_time,
            end_time=end_time,
            max_capacity=max_capacity,
            safe_threshold=safe_threshold,
            warning_threshold=warning_threshold,
            event_type=event_type,
            is_paid=is_paid,
            price=price,
            is_group_event=is_group_event,
            group_size=group_size,
            host=current_user if current_user.is_authenticated and current_user.role in {"host", "admin", "faculty_coordinator"} else None,
        )
        db.session.add(event)
        db.session.commit()
        flash("Event created successfully.", "success")
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        flash(str(exc), "danger")

    return redirect(url_for("index"))


@app.route("/events/<int:event_id>/edit", methods=["GET", "POST"])
@login_required
@require_role("host", "admin", "faculty_coordinator")
def edit_event(event_id: int):
    event = Event.query.get_or_404(event_id)
    
    # Check authorization: Admin/Faculty can edit all, Host only their own
    if current_user.role == "host" and event.host_id != current_user.id:
        flash("You are not authorized to edit this event.", "danger")
        return redirect(url_for("host_dashboard"))

    if request.method == "POST":
        try:
            event.name = request.form.get("name", "").strip()
            event.location = request.form.get("location", "").strip()
            start_time_str = request.form.get("start_time")
            end_time_str = request.form.get("end_time")
            event.max_capacity = int(request.form.get("max_capacity") or 0)
            event.safe_threshold = float(request.form.get("safe_threshold") or 70) / 100.0
            event.warning_threshold = float(request.form.get("warning_threshold") or 90) / 100.0
            event.event_type = request.form.get("event_type") or "general"
            event.is_paid = request.form.get("is_paid") == "1"
            event.price = float(request.form.get("price") or 0.0) if event.is_paid else 0.0
            event.is_group_event = request.form.get("is_group_event") == "1"
            event.group_size = int(request.form.get("group_size") or 1)

            if not event.name or not event.location or event.max_capacity <= 0:
                raise ValueError("Please fill all fields and use a positive capacity.")

            event.start_time = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M")
            event.end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M")

            if event.end_time <= event.start_time:
                raise ValueError("End time must be after start time.")

            if not (0 < event.safe_threshold < event.warning_threshold <= 1):
                raise ValueError("Thresholds must be between 0–100 and safe < warning.")

            db.session.commit()
            flash("Event updated successfully.", "success")
            return redirect(url_for("host_dashboard") if current_user.role == "host" else url_for("admin_dashboard"))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")

    return render_template("edit_event.html", event=event)


@app.route("/events/<int:event_id>/block", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator")
def block_event_entry(event_id: int):
    event = Event.query.get_or_404(event_id)
    event.entry_blocked = True
    db.session.commit()
    flash("Entry blocked for this event.", "warning")
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/events/<int:event_id>/unblock", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator")
def unblock_event_entry(event_id: int):
    event = Event.query.get_or_404(event_id)
    event.entry_blocked = False
    db.session.commit()
    flash("Entry unblocked for this event.", "success")
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/events/<int:event_id>/thresholds", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator", "host")
def update_event_thresholds(event_id: int):
    """Update safe and warning thresholds for an event (values in % from form)."""
    event = Event.query.get_or_404(event_id)

    try:
        safe_pct = float(request.form.get("safe_threshold") or 70)
        warn_pct = float(request.form.get("warning_threshold") or 90)
        safe = safe_pct / 100.0
        warn = warn_pct / 100.0

        if not (0 < safe < warn <= 1):
            raise ValueError("Thresholds must be between 0–100 and safe < warning.")

        event.safe_threshold = safe
        event.warning_threshold = warn
        db.session.commit()
        flash("Thresholds updated successfully.", "success")
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        flash(str(exc), "danger")

    return redirect(request.referrer or url_for("host_dashboard"))


@app.route("/events/<int:event_id>/poster", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator", "host")
def upload_event_poster(event_id: int):
    event = Event.query.get_or_404(event_id)

    file = request.files.get("poster")
    if not file or file.filename == "":
        flash("Please choose an image file to upload.", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    _, ext = os.path.splitext(file.filename)
    ext = ext.lower()
    allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".mp4", ".webm"}
    if ext not in allowed_exts:
        flash("Only JPG, PNG, WEBP, GIF, MP4, or WEBM files are allowed.", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    folder = os.path.join(app.root_path, "static", "event_posters")
    os.makedirs(folder, exist_ok=True)
    filename = f"event_{event.id}{ext}"
    filepath = os.path.join(folder, filename)
    file.save(filepath)

    event.poster_filename = filename
    db.session.commit()

    flash("Event poster updated.", "success")
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/events/<int:event_id>/zones/create", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator")
def create_zone(event_id: int):
    event = Event.query.get_or_404(event_id)
    name = request.form.get("name", "").strip()
    max_capacity = int(request.form.get("max_capacity") or 0)
    
    if not name or max_capacity <= 0:
        flash("Invalid zone data.", "danger")
        return redirect(url_for("event_detail", event_id=event_id))
    
    zone = Zone(name=name, max_capacity=max_capacity, event_id=event_id)
    db.session.add(zone)
    db.session.commit()
    flash("Zone created successfully.", "success")
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/events/<int:event_id>/gates/create", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator")
def create_gate(event_id: int):
    event = Event.query.get_or_404(event_id)
    name = request.form.get("name", "").strip()
    
    if not name:
        flash("Gate name is required.", "danger")
        return redirect(url_for("event_detail", event_id=event_id))
    
    gate = Gate(name=name, event_id=event_id)
    db.session.add(gate)
    db.session.commit()
    flash("Gate created successfully.", "success")
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/events/<int:event_id>/time-slots/create", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator")
def create_time_slot(event_id: int):
    event = Event.query.get_or_404(event_id)
    start_time_str = request.form.get("start_time")
    end_time_str = request.form.get("end_time")
    max_participants = int(request.form.get("max_participants") or 0)
    
    try:
        start_time = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M")
        end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M")
        
        if end_time <= start_time or max_participants <= 0:
            raise ValueError("Invalid time slot data.")
        
        time_slot = TimeSlot(
            start_time=start_time,
            end_time=end_time,
            max_participants=max_participants,
            event_id=event_id
        )
        db.session.add(time_slot)
        db.session.commit()
        flash("Time slot created successfully.", "success")
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        flash(str(exc), "danger")
    
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/announcements/create", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator")
def create_announcement():
    title = request.form.get("title", "").strip()
    message = request.form.get("message", "").strip()
    event_id = request.form.get("event_id")
    
    if not title or not message:
        flash("Title and message are required.", "danger")
        return redirect(request.referrer or url_for("index"))
    
    announcement = Announcement(
        title=title,
        message=message,
        event_id=int(event_id) if event_id else None,
        created_by_id=current_user.id,
        is_active=True
    )
    db.session.add(announcement)
    db.session.commit()
    flash("Announcement created successfully.", "success")
    return redirect(request.referrer or url_for("index"))


@app.route("/announcements/<int:announcement_id>/toggle", methods=["POST"])
@login_required
@require_role("admin", "faculty_coordinator")
def toggle_announcement(announcement_id: int):
    announcement = Announcement.query.get_or_404(announcement_id)
    announcement.is_active = not announcement.is_active
    db.session.commit()
    flash("Announcement updated.", "success")
    return redirect(request.referrer or url_for("index"))


@app.route("/api/events/<int:event_id>/students")
@login_required
def api_get_enrolled_students(event_id: int):
    """API endpoint to get students enrolled in a specific event."""
    tickets = Ticket.query.filter_by(event_id=event_id).all()
    students = [
        {"id": t.student.id, "name": t.student.name, "roll_number": t.student.roll_number or t.student.email}
        for t in tickets
    ]
    return jsonify(students)


@app.route("/api/recent_logs")
@login_required
def api_recent_logs():
    """API endpoint to get recent check-in logs for the current user."""
    event_id = request.args.get("event_id")
    
    query = Ticket.query.filter(Ticket.checked_in == True)
    
    if event_id:
        query = query.filter(Ticket.event_id == int(event_id))

    if current_user.role == "admin":
        logs = query.order_by(Ticket.checked_in_at.desc()).limit(10).all()
    elif current_user.role == "host":
        logs = query.join(Event).filter(Event.host_id == current_user.id).order_by(Ticket.checked_in_at.desc()).limit(10).all()
    elif current_user.role == "volunteer":
        logs = query.order_by(Ticket.checked_in_at.desc()).limit(10).all()
    else:
        return jsonify({"error": "Not authorized"}), 403

    result = [
        {
            "time": l.checked_in_at.strftime('%H:%M:%S'),
            "student_name": l.student.name,
            "event_name": l.event.name,
            "event_id": l.event_id,
            "status": "success"
        }
        for l in logs
    ]
    return jsonify(result)


@app.route("/api/events/<int:event_id>")
def api_get_event(event_id: int):
    event = Event.query.get_or_404(event_id)
    return jsonify(_serialize_event(event))


@app.route("/api/events/<int:event_id>/update_count", methods=["POST"])
@login_required
def api_update_count(event_id: int):
    event = Event.query.get_or_404(event_id)
    if current_user.role not in {"admin", "faculty_coordinator", "host"}:
        return jsonify({"error": "Not authorized"}), 403
    
    data = request.get_json(silent=True) or {}
    delta = int(data.get("delta") or 0)

    new_count = event.current_count + delta
    new_count = max(0, min(new_count, event.max_capacity))
    event.current_count = new_count
    db.session.commit()

    return jsonify(_serialize_event(event))


@app.route("/events/<int:event_id>/tickets/create", methods=["POST"])
@login_required
def create_ticket(event_id: int):
    event = Event.query.get_or_404(event_id)
    
    if event.entry_blocked:
        flash("Entry is currently blocked for this event.", "danger")
        return redirect(url_for("event_detail", event_id=event.id))
    
    if current_user.role != "student":
        flash("Only students can register for passes.", "danger")
        return redirect(url_for("event_detail", event_id=event.id))

    # Check event capacity
    if event.current_count >= event.max_capacity:
        flash("This event is already at full capacity.", "danger")
        return redirect(url_for("event_detail", event_id=event.id))

    quantity = int(request.form.get("quantity") or 1)
    existing_count = Ticket.query.filter_by(event_id=event.id, student_id=current_user.id).count()
    
    max_allowed = event.group_size if event.is_group_event else 4
    if existing_count + quantity > max_allowed:
        flash(f"You can only have a maximum of {max_allowed} passes for this event. You already have {existing_count}.", "warning")
        return redirect(url_for("event_detail", event_id=event.id))

    time_slot_id = request.form.get("time_slot_id")
    time_slot = None
    if time_slot_id:
        time_slot = TimeSlot.query.get(int(time_slot_id))
        if not time_slot or time_slot.event_id != event.id:
            flash("Invalid time slot selected.", "danger")
            return redirect(url_for("event_detail", event_id=event.id))
        if time_slot.is_full:
            flash("This time slot is full.", "danger")
            return redirect(url_for("event_detail", event_id=event.id))

    team_name = request.form.get("team_name")
    member_details = request.form.get("member_details")
    
    # Capture individual member names if provided
    member_names = request.form.getlist("member_names[]")

    created_tickets = []
    for i in range(quantity):
        if time_slot:
            time_slot.current_participants += 1
            
        qr_token = secrets.token_hex(16)
        
        # Determine the name for this specific QR code
        specific_guest_name = None
        if event.is_group_event:
            if i < len(member_names) and member_names[i].strip():
                specific_guest_name = member_names[i].strip()
            else:
                specific_guest_name = f"Member {i+1}"
        
        ticket = Ticket(
            qr_token=qr_token,
            event=event,
            student=current_user,
            paid=not event.is_paid,
            time_slot=time_slot,
            team_name=team_name if event.is_group_event else None,
            member_details=member_details if event.is_group_event else None,
            guest_name=specific_guest_name
        )
        db.session.add(ticket)
        created_tickets.append(ticket)
    
    db.session.commit()

    if event.is_paid:
        ticket_ids = ",".join([str(t.id) for t in created_tickets])
        flash(f"Successfully created {quantity} passes. Complete payment to activate them.", "success")
        return redirect(url_for("bulk_payment_page", ticket_ids=ticket_ids))
    else:
        flash(f"Registration successful! {quantity} free passes are now active.", "success")
        return redirect(url_for("view_pass", ticket_id=created_tickets[0].id))


@app.route("/my-tickets")
@login_required
def my_tickets():
    if current_user.role != "student":
        flash("Only students can view their tickets.", "danger")
        return redirect(url_for("index"))

    now = datetime.now()
    tickets = Ticket.query.filter_by(student_id=current_user.id).order_by(Ticket.created_at.desc()).all()
    # Enrolled event IDs
    enrolled_event_ids = [t.event_id for t in tickets]
    
    # Get all coupons for the student
    all_student_coupons = FoodCoupon.query.filter_by(student_id=current_user.id).order_by(FoodCoupon.created_at.desc()).all()
    
    # Filter coupons: 
    # 1. General coupons (no event_id)
    # 2. Event coupons only if student is enrolled in that event
    coupons = [
        c for c in all_student_coupons 
        if c.event_id is None or c.event_id in enrolled_event_ids
    ]
    
    return render_template("my_tickets.html", tickets=tickets, coupons=coupons, now=now)


@app.route("/tickets/<int:ticket_id>/pay", methods=["GET", "POST"])
@login_required
def payment_page(ticket_id: int):
    # Redirect to bulk payment for consistency
    return redirect(url_for("bulk_payment_page", ticket_ids=str(ticket_id)))


@app.route("/bulk-payment/<ticket_ids>", methods=["GET", "POST"])
@login_required
def bulk_payment_page(ticket_ids: str):
    id_list = [int(tid) for tid in ticket_ids.split(",")]
    tickets = Ticket.query.filter(Ticket.id.in_(id_list)).all()
    
    if not tickets:
        flash("No valid tickets found for payment.", "danger")
        return redirect(url_for("my_tickets"))

    # Security check: all tickets must belong to current user (or admin)
    for t in tickets:
        if t.student_id != current_user.id and current_user.role != "admin":
            flash("Not authorized to view this payment.", "danger")
            return redirect(url_for("index"))

    # Calculate total
    # If it's a group event, we charge the event price once for the entire group
    # Otherwise, we sum the price for each individual ticket
    first_ticket = tickets[0]
    if first_ticket.event.is_group_event:
        total_price = first_ticket.event.price
    else:
        total_price = sum([t.event.price for t in tickets if t.event.is_paid])

    if request.method == "POST":
        for t in tickets:
            t.paid = True
        db.session.commit()
        flash(f"Payment of ₹{total_price:.2f} successful. {len(tickets)} passes are now active.", "success")
        return redirect(url_for("my_tickets"))

    return render_template("payment.html", tickets=tickets, total_price=total_price)


def _ensure_qr_folder():
    folder = app.config["QR_FOLDER"]
    os.makedirs(folder, exist_ok=True)
    return folder


@app.route("/tickets/<int:ticket_id>/pass")
@login_required
def view_pass(ticket_id: int):
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.student_id != current_user.id and current_user.role not in {"host", "admin", "faculty_coordinator"}:
        flash("Not authorized to view this pass.", "danger")
        return redirect(url_for("index"))

    if not ticket.paid:
        flash("Complete payment before using the pass.", "warning")
        return redirect(url_for("payment_page", ticket_id=ticket.id))

    # Security: QR code now only contains the raw token, not a URL.
    # This prevents generic phone scanners from automatically verifying the pass.
    qr_data = ticket.qr_token
    qr_folder = _ensure_qr_folder()
    qr_filename = f"ticket_{ticket.id}.png"
    qr_path = os.path.join(qr_folder, qr_filename)

    if not os.path.exists(qr_path):
        img = qrcode.make(qr_data)
        img.save(qr_path)

    qr_url = url_for("static", filename=f"qr/{qr_filename}")
    
    # Get specific food coupon linked to this pass
    linked_coupon = FoodCoupon.query.filter_by(ticket_id=ticket.id).first()
    
    # If no specific linked coupon, maybe show general student coupons for this event
    if linked_coupon:
        coupons = [linked_coupon]
    else:
        coupons = FoodCoupon.query.filter_by(event_id=ticket.event_id, student_id=ticket.student_id, ticket_id=None).all()
    
    return render_template("ticket_pass.html", ticket=ticket, qr_url=qr_url, coupons=coupons)


@app.route("/api/verify/<path:token>")
def api_verify_qr(token: str):
    """API endpoint for volunteer scanner"""
    # Just in case the token has multiple segments (e.g., from an old URL-based scan)
    if "/" in token:
        token = token.strip("/").split("/")[-1]
    
    token = token.strip()
    print(f"DEBUG: Verifying token: '{token}'")
    ticket = Ticket.query.filter_by(qr_token=token).first()
    if not ticket:
        print(f"DEBUG: Token '{token}' not found in database.")
        return jsonify({"status": "invalid", "message": f"Invalid pass: {token}"}), 404

    event = ticket.event
    
    if event.entry_blocked:
        return jsonify({"status": "blocked", "message": "Entry blocked"}), 403
    
    if ticket.checked_in:
        return jsonify({"status": "already_used", "message": "Pass already used"}), 400

    if ticket.time_slot:
        now = datetime.now()
        if now < ticket.time_slot.start_time:
            return jsonify({"status": "early", "message": f"Time slot starts at {ticket.time_slot.start_time.strftime('%H:%M')}"}), 400
        if now > ticket.time_slot.end_time:
            return jsonify({"status": "expired", "message": "Time slot expired"}), 400

    ticket.checked_in = True
    ticket.checked_in_at = datetime.now()
    
    # Always increment if it's a new check-in
    event.current_count += 1
    
    # Update gate if provided
    gate_id = request.args.get("gate_id")
    if gate_id:
        gate = Gate.query.get(gate_id)
        if gate and gate.event_id == event.id:
            gate.entry_count += 1
            ticket.gate_id = gate.id
            
    # Update zone if provided
    zone_id = request.args.get("zone_id")
    if zone_id:
        zone = Zone.query.get(zone_id)
        if zone and zone.event_id == event.id:
            zone.current_count += 1
    
    db.session.commit()

    return jsonify({
        "status": "ok",
        "message": "Verified",
        "student_name": ticket.student.name,
        "event_name": event.name,
        "event_id": event.id,
    })


def _serialize_event(event: Event) -> dict:
    zones_data = [{"id": z.id, "name": z.name, "current_count": z.current_count, "max_capacity": z.max_capacity, "occupancy_ratio": round(z.occupancy_ratio, 3)} for z in event.zones]
    gates_data = [{"id": g.id, "name": g.name, "entry_count": g.entry_count, "exit_count": g.exit_count, "net_count": g.net_count} for g in event.gates]
    
    return {
        "id": event.id,
        "name": event.name,
        "location": event.location,
        "start_time": event.start_time.isoformat(),
        "end_time": event.end_time.isoformat(),
        "max_capacity": event.max_capacity,
        "current_count": event.current_count,
        "occupancy_ratio": round(event.occupancy_ratio, 3),
        "status": event.status,
        "status_message": event.status_message,
        "occupancy_percent": int(event.occupancy_ratio * 100),
        "entry_blocked": event.entry_blocked,
        "zones": zones_data,
        "gates": gates_data,
    }


def _generate_pdf(title, headers, data):
    """Helper to generate a PDF with a table."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 12))

    # Table
    table_data = [headers] + data
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


@app.route("/admin/export_users_pdf")
@login_required
@require_role("admin")
def export_users_pdf():
    users = User.query.all()
    data = [[u.id, u.name, u.email, u.role, u.roll_number or "N/A"] for u in users]
    headers = ["ID", "Name", "Email", "Role", "Roll Number"]
    buffer = _generate_pdf("User Registration Details", headers, data)
    return send_file(buffer, as_attachment=True, download_name="user_registrations.pdf", mimetype="application/pdf")


@app.route("/host/remove_user/<int:user_id>", methods=["POST"])
@login_required
@require_role("host", "admin")
def remove_user(user_id: int):
    user = User.query.get_or_404(user_id)
    
    # Security: Host can only remove faculty or volunteers
    if user.role not in {"faculty_coordinator", "volunteer"}:
        flash("You can only remove faculty coordinators or volunteers.", "danger")
        return redirect(url_for("host_dashboard"))
    
    # Prevent host from removing themselves or other hosts/admins
    if user.id == current_user.id:
        flash("You cannot remove yourself.", "danger")
        return redirect(url_for("host_dashboard"))

    try:
        db.session.delete(user)
        db.session.commit()
        flash(f"User {user.name} has been removed successfully.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Error removing user: {str(exc)}", "danger")

    return redirect(url_for("host_dashboard"))


@app.route("/host/export_entry_logs_excel")
@login_required
@require_role("host", "admin")
def export_entry_logs_excel():
    if current_user.role == "admin":
        logs = Ticket.query.filter(Ticket.checked_in == True).order_by(Ticket.checked_in_at.desc()).all()
    else:
        logs = Ticket.query.join(Event).filter(Event.host_id == current_user.id, Ticket.checked_in == True).order_by(Ticket.checked_in_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entry Logs"
    
    headers = ["Student Name", "Roll Number", "Event Name", "Check-in Time", "Gate"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for l in logs:
        ws.append([
            l.student.name,
            l.student.roll_number or l.student.email,
            l.event.name,
            l.checked_in_at.strftime('%Y-%m-%d %H:%M'),
            l.gate.name if l.gate else "N/A"
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="entry_logs.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/host/export_entry_logs_pdf")
@login_required
@require_role("host", "admin")
def export_entry_logs_pdf():
    if current_user.role == "admin":
        logs = Ticket.query.filter(Ticket.checked_in == True).order_by(Ticket.checked_in_at.desc()).all()
    else:
        logs = Ticket.query.join(Event).filter(Event.host_id == current_user.id, Ticket.checked_in == True).order_by(Ticket.checked_in_at.desc()).all()
    
    data = [[l.student.name, l.event.name, l.checked_in_at.strftime('%Y-%m-%d %H:%M'), l.gate.name if l.gate else "N/A"] for l in logs]
    headers = ["Student", "Event", "Check-in Time", "Gate"]
    buffer = _generate_pdf("Entry Logs Report", headers, data)
    return send_file(buffer, as_attachment=True, download_name="entry_logs.pdf", mimetype="application/pdf")


@app.route("/host/export_enrollments_pdf/<int:event_id>")
@login_required
@require_role("host", "admin")
def export_enrollments_pdf(event_id):
    event = Event.query.get_or_404(event_id)
    if current_user.role != "admin" and event.host_id != current_user.id:
        flash("Unauthorized", "danger")
        return redirect(url_for("host_dashboard"))

    tickets = Ticket.query.filter_by(event_id=event_id).all()
    data = []
    for t in tickets:
        student_name = t.student.name if t.student else "Unknown"
        specific_name = t.guest_name if t.guest_name else student_name
        team_name = t.team_name or "N/A"
        data.append([specific_name, student_name, team_name, t.student.roll_number or "N/A", "Yes" if t.paid else "No"])
        
    headers = ["Member Name", "Registered By", "Team", "Roll Number", "Paid"]
    buffer = _generate_pdf(f"Enrollments for {event.name}", headers, data)
    return send_file(buffer, as_attachment=True, download_name=f"enrollments_{event_id}.pdf", mimetype="application/pdf")


@app.route("/host/export_food_history_pdf")
@login_required
@require_role("host", "admin")
def export_host_food_history_pdf():
    if current_user.role == "admin":
        redemptions = FoodCouponRedemption.query.order_by(FoodCouponRedemption.redeemed_at.desc()).all()
    else:
        redemptions = FoodCouponRedemption.query.join(Event).filter(Event.host_id == current_user.id).order_by(FoodCouponRedemption.redeemed_at.desc()).all()
    
    data = []
    for r in redemptions:
        student_name = r.coupon.student.name if r.coupon and r.coupon.student else "Unknown"
        event_name = r.event.name if r.event else "General"
        food_item = r.coupon.food_item if r.coupon else "N/A"
        data.append([r.redeemed_at.strftime('%Y-%m-%d %H:%M'), student_name, event_name, food_item])
        
    headers = ["Redemption Time", "Student", "Event", "Food Item"]
    buffer = _generate_pdf("Host Food Coupon History", headers, data)
    return send_file(buffer, as_attachment=True, download_name="host_food_history.pdf", mimetype="application/pdf")


@app.route("/host/export_food_history_excel")
@login_required
@require_role("host", "admin")
def export_host_food_history_excel():
    if current_user.role == "admin":
        redemptions = FoodCouponRedemption.query.order_by(FoodCouponRedemption.redeemed_at.desc()).all()
    else:
        redemptions = FoodCouponRedemption.query.join(Event).filter(Event.host_id == current_user.id).order_by(FoodCouponRedemption.redeemed_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Food Redemption History"
    
    # Headers
    headers = ["Redemption Time", "Student Name", "Event Name", "Food Item", "Quantity", "Redeemed By"]
    ws.append(headers)
    
    # Styling headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for r in redemptions:
        student_name = r.coupon.student.name if r.coupon and r.coupon.student else "Unknown"
        event_name = r.event.name if r.event else "General"
        food_item = r.coupon.food_item if r.coupon else "N/A"
        quantity = r.coupon.quantity if r.coupon else 1
        redeemed_by = r.redeemed_by.name if r.redeemed_by else "System"
        ws.append([r.redeemed_at.strftime('%Y-%m-%d %H:%M'), student_name, event_name, food_item, quantity, redeemed_by])

    # Column width adjustment
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="host_food_history.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/host/export_enrollment_history_excel")
@login_required
@require_role("host", "admin")
def export_enrollment_history_excel():
    if current_user.role == "admin":
        tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()
    else:
        tickets = Ticket.query.join(Event).filter(Event.host_id == current_user.id).order_by(Ticket.created_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pass Generation History"
    
    # Headers
    headers = ["Registration Time", "Member Name", "Registered By", "Team", "Roll Number", "Event Name", "Payment Status", "Checked In", "Check-in Time"]
    ws.append(headers)
    
    # Styling headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for t in tickets:
        student_name = t.student.name if t.student else "Unknown"
        specific_name = t.guest_name if t.guest_name else student_name
        team_name = t.team_name or "Individual"
        roll_number = t.student.roll_number or t.student.email if t.student else "N/A"
        event_name = t.event.name if t.event else "Unknown"
        payment_status = "Paid" if t.paid else "Pending"
        checked_in = "Yes" if t.checked_in else "No"
        checkin_time = t.checked_in_at.strftime('%Y-%m-%d %H:%M') if t.checked_in_at else "N/A"
        
        ws.append([
            t.created_at.strftime('%Y-%m-%d %H:%M'),
            specific_name,
            student_name,
            team_name,
            roll_number,
            event_name,
            payment_status,
            checked_in,
            checkin_time
        ])

    # Column width adjustment
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="pass_generation_history.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/canteen/export_coupon_history_pdf")
@login_required
@require_role("canteen_owner", "admin")
def export_coupon_history_pdf():
    redemptions = FoodCouponRedemption.query.order_by(FoodCouponRedemption.redeemed_at.desc()).all()
    data = [[r.redeemed_at.strftime('%Y-%m-%d %H:%M'), r.redeemed_by.name if r.redeemed_by else "System", r.qr_data[:20] + "..."] for r in redemptions]
    headers = ["Redemption Time", "Redeemed By", "QR Data Snippet"]
    buffer = _generate_pdf("Food Coupon Redemption History", headers, data)
    return send_file(buffer, as_attachment=True, download_name="coupon_history.pdf", mimetype="application/pdf")


if __name__ == "__main__":
    with app.app_context():
        # For development: Ensure QR folder exists
        os.makedirs(app.config["QR_FOLDER"], exist_ok=True)
        
        # Check if tables exist by trying to query User
        try:
            User.query.first()
        except Exception:
            print("Creating database tables...")
            db.create_all()
        
        # Create a default admin user for testing
        admin = User.query.filter_by(email="admin@college.edu").first()
        if not admin:
            admin = User(name="Admin", email="admin@college.edu", role="admin")
            admin.set_password("admin123")
            db.session.add(admin)
            db.session.commit()
            print("Default admin created: admin@college.edu / admin123")

        # Demo users for testing different roles
        host_user = User.query.filter_by(email="host@college.edu").first()
        if not host_user:
            host_user = User(name="Host User", email="host@college.edu", role="host")
            host_user.set_password("host123")
            db.session.add(host_user)
            db.session.commit()
            print("Demo host created: host@college.edu / host123")

        canteen_user = User.query.filter_by(email="canteen@college.edu").first()
        if not canteen_user:
            canteen_user = User(name="Canteen Owner", email="canteen@college.edu", role="canteen_owner")
            canteen_user.set_password("canteen123")
            db.session.add(canteen_user)
            db.session.commit()
            print("Demo canteen created: canteen@college.edu / canteen123")

        security_user = User.query.filter_by(email="security@college.edu").first()
        if not security_user:
            security_user = User(name="Security Officer", email="security@college.edu", role="security")
            security_user.set_password("security123")
            db.session.add(security_user)
            db.session.commit()
            print("Demo security created: security@college.edu / security123")

        volunteer_user = User.query.filter_by(email="volunteer@college.edu").first()
        if not volunteer_user:
            volunteer_user = User(name="Volunteer User", email="volunteer@college.edu", role="volunteer")
            volunteer_user.set_password("volunteer123")
            db.session.add(volunteer_user)
            db.session.commit()
            print("Demo volunteer created: volunteer@college.edu / volunteer123")

        faculty_user = User.query.filter_by(email="faculty@college.edu").first()
        if not faculty_user:
            faculty_user = User(name="Faculty Coordinator", email="faculty@college.edu", role="faculty_coordinator")
            faculty_user.set_password("faculty123")
            db.session.add(faculty_user)
            db.session.commit()
            print("Demo faculty created: faculty@college.edu / faculty123")

        student1 = User.query.filter_by(email="student1@college.edu").first()
        if not student1:
            student1 = User(name="Student One", email="student1@college.edu", role="student")
            student1.set_password("student123")
            db.session.add(student1)
            db.session.commit()
            print("Demo student created: student1@college.edu / student123")

        student2 = User.query.filter_by(email="student2@college.edu").first()
        if not student2:
            student2 = User(name="Student Two", email="student2@college.edu", role="student")
            student2.set_password("student123")
            db.session.add(student2)
            db.session.commit()
            print("Demo student created: student2@college.edu / student123")

        # Demo events for dashboards
        if Event.query.count() == 0:
            now = datetime.now()
            tech_fest = Event(
                name="Tech Fest 2026",
                location="Main Auditorium",
                start_time=now + timedelta(days=1),
                end_time=now + timedelta(days=1, hours=4),
                max_capacity=500,
                safe_threshold=0.7,
                warning_threshold=0.9,
                event_type="technical",
                host=host_user,
            )

            cultural_night = Event(
                name="Cultural Night",
                location="Open Air Theatre",
                start_time=now + timedelta(days=2),
                end_time=now + timedelta(days=2, hours=3),
                max_capacity=800,
                safe_threshold=0.7,
                warning_threshold=0.9,
                event_type="cultural",
                host=host_user,
            )

            seminar = Event(
                name="AI & Data Science Seminar",
                location="Seminar Hall B",
                start_time=now + timedelta(days=3),
                end_time=now + timedelta(days=3, hours=2),
                max_capacity=300,
                safe_threshold=0.7,
                warning_threshold=0.9,
                event_type="seminar",
                host=host_user,
            )

            db.session.add_all([tech_fest, cultural_night, seminar])
            db.session.commit()

            # Add Gates to Tech Fest
            gate_main = Gate(name="Main Entrance", event_id=tech_fest.id)
            gate_side = Gate(name="Side Entry", event_id=tech_fest.id)
            
            # Add Zones to Tech Fest
            zone_stage = Zone(name="Stage Area", max_capacity=100, event_id=tech_fest.id)
            zone_seating = Zone(name="Seating Area", max_capacity=400, event_id=tech_fest.id)

            db.session.add_all([gate_main, gate_side, zone_stage, zone_seating])
            db.session.commit()
            print("Demo events and gates/zones created for dashboards.")
        
        import socket
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
        print(f"\n" + "="*50)
        print(f"MOBILE ACCESS INFO")
        print(f"Connect your phone to the SAME WiFi as this PC.")
        print(f"Open this URL on your mobile: http://{local_ip}:8080")
        print("="*50 + "\n")

    app.run(host="0.0.0.0", port=8080, debug=True)
